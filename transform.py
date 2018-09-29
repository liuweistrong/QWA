from load import get_library
from openpyxl import load_workbook
from PATH import *
#书写一下转化的函数

#生成待录取的精简列表
def get_easyList(pin): #返回精简的带录取列表
    appli_full_info = get_library(1)
    job_book = load_workbook(JOBINGO_PATH)
    ws = job_book['worker']
    easyList = []
    max_row = ws.max_row
    for row in range(2,max_row+1):
        num = ws.cell(row,1).value
        for appli in appli_full_info:
            if appli[0] == num:
                #构造精简列表
                appli_easy_list = []
                appli_easy_list.append(appli[0])#编号
                appli_easy_list.append(appli[6]) #贫困1
                appli_easy_list.append(appli[7])#服从调剂

                appli_easy_list.append(appli[8])#岗位一
                appli_easy_list.append(appli[9])#岗位二
                appli_easy_list.append(appli[10])# 岗位三

                easyList.append(appli_easy_list)

    #划分easyList
    easyList_pinkun = []
    easyList_feipin = []
    for easy in easyList:
        if easy[1] == '是':
            del easy[1]
            easyList_pinkun.append(easy)
        else:
            del easy[1]
            easyList_feipin.append(easy)
    if pin:
        return easyList_pinkun
    else:
        return easyList_feipin

def enrolled2full(enrolled_list):
    full_info = get_library(1)
    enrolled_full=[]
    if(enrolled_list):
        for enroll_dict in enrolled_list:
            for bianhao in enroll_dict.keys(): #就一个，取出来编号
                for appli in full_info:
                    if appli[0] == bianhao:
                        enroller = appli[:6]
                        enroller.append(enroll_dict[bianhao])
                        enrolled_full.append(enroller)
    return enrolled_full
