import os
import time,datetime
import shutil
from openpyxl import Workbook
from openpyxl import load_workbook

#ROOT_DIR = C:\Users\liuweistrong\Documents\TWA
LIBRARY_PATH = r'C:\Users\liuweistrong\Documents\TWA\work\library.xlsx'
JOBINGO_PATH = r'C:\Users\liuweistrong\Documents\TWA\work\willGetJob.xlsx'
RESULT_PATH = r'C:\Users\liuweistrong\Documents\TWA\result.xlsx'

    # 1 生成待录取精简列表
    # appli_lite_pinkun = [[15，True，10000，10002，12225],[],[]...] [编号，服从调剂，岗位编号1，2，3]
    # appli_lite_feipin
    #0.1版本只读取A+


#load library
def get_library(flag):
    appli_book = load_workbook(LIBRARY_PATH)
    library = []
    if flag == 1:  # 获取所有报名信息到列表
        ws = appli_book['library']
        max_row = ws.max_row
        #max col固定是14
        for row in range(2,max_row+1):
            appli_info = []
            for j in range(1,14+1):
                appli_info.append(ws.cell(row,j).value)
            library.append(appli_info)

            # [1, '姚佳莹', '女', 3017225028, '经管学院', 17302200916, '否', '否', '000001', None, None, '勤管办义务岗主管',None, None]
        return library

    if flag == 2:  #获取岗位需求信息到字典
        ws = appli_book['jobInfo']
        max_row = ws.max_row
        # max col固定是4
        for row in range(2, max_row + 1):
            job_info = {}
            job_info['num']=ws.cell(row, 1).value#编号
            job_info['pinkun']=ws.cell(row,3).value
            job_info['feipin']=ws.cell(row,4).value
            library.append(job_info)


            # class 'list'>: [{'num': '000001', 'pinkun': 3, 'feipin': 0}, {'num': '000002', 'pinkun': 2, 'feipin': 1},
            #                 {'num': '000004', 'pinkun': 4, 'feipin': 1}, {'num': '210101', 'pinkun': 1, 'feipin': None},
            #                 {'num': '280301', 'pinkun': 1, 'feipin': None},

        return library

    #def 从编号获取所有信息
def num2allinfo(num):
    pass
#生成待录取的精简列表
def get_easyList(pin): #返回精简的带录取列表
    appli_full_info = get_library(1)
    job_book = load_workbook(JOBINGO_PATH)
    ws = job_book['worker']
    easyList = []
    max_row = ws.max_row
    for row in range(2,max_row+1):
        num = ws.cell(row,1).value
        for appli in appli_full_info:
            if appli[0] == num:
                #构造精简列表
                appli_easy_list = []
                appli_easy_list.append(appli[0])#编号
                appli_easy_list.append(appli[6]) #贫困1
                appli_easy_list.append(appli[7])#服从调剂

                appli_easy_list.append(appli[8])#岗位一
                appli_easy_list.append(appli[9])#岗位二
                appli_easy_list.append(appli[10])# 岗位三

                easyList.append(appli_easy_list)

    #划分easyList
    easyList_pinkun = []
    easyList_feipin = []
    for easy in easyList:
        if easy[1] == '是':
            del easy[1]
            easyList_pinkun.append(easy)
        else:
            del easy[1]
            easyList_feipin.append(easy)
    if pin:
        return easyList_pinkun
    else:
        return easyList_feipin

#对于某一个岗位的录取方法
def exeEnroll(jobNum,pin,easy_list):
    if not pin:

        for appli in easy_list:
            #第一志愿查找
            if appli[2] == jobNum:
                return appli[0]
        for appli in easy_list:
            #第二志愿查找
            if appli[3] == jobNum:
                return appli[0]
        for appli in easy_list:
            #第三志愿查找
            if appli[4] == jobNum:
                return appli[0]
        for appli in easy_list:
            #检索服从调剂的同学
            if appli[1] == '是':
                return appli[0]
        return None
    if pin:

        for appli in easy_list:
            #第一志愿查找
            if appli[2] == jobNum:
                return appli[0]
        for appli in easy_list:
            #第二志愿查找
            if appli[3] == jobNum:
                return appli[0]
        for appli in easy_list:
            #第三志愿查找
            if appli[4] == jobNum:
                return appli[0]
        for appli in easy_list:
            #检索服从调剂的同学
            if appli[1] == '是':
                return appli[0]
        return None


def enroll_feipin(jobInfo):
    print('首先我们来进行非贫困生的录取')
    easy_list = get_easyList(False)
    print('获得待录取资格的非贫困生人数是 {0}'.format(len(easy_list)))
    enrolled = []
    count = 0
    for jobDict in jobInfo:
        n = jobDict.get('feipin')
        if n != 0 and n!= None:
            count = count + n
    if count:
        print('非贫困生录取名额总数是 {0}'.format(count))
        for jobDict in jobInfo: #对招聘信息中招收非贫困生的岗位进行遍历
            if jobDict.get('feipin') != 0 and jobDict.get('feipin') != None:
                n = jobDict['feipin']
                jobNum = jobDict.get('num')
                for _ in range(n): # 找几个人就循环多少次
                    bianhao = exeEnroll(jobNum,pin=False,easy_list=easy_list)
                    if bianhao != None: #录取了某个同学
                        print('  恭喜编号是{0}的同学(非贫困生)被录取岗位{1}'.format(bianhao,jobNum))
                        #操作录取的列表改变  改变jobinfo 增加录取的列表 减少带录取的easylist
                        enrollItem = {}
                        enrollItem[bianhao] = jobNum
                        enrolled.append(enrollItem)

                        jobDict['feipin'] -= 1

                        i = 0
                        for appli in easy_list:
                            if appli[0] == bianhao:
                                del easy_list[i]
                            i+=1
                    else:
                        jobDict['pinkun'] += jobDict['feipin']
                        msg = '  因为非贫困生没能招满名额，对于岗位{0}，将非贫困生剩余的{1}个名额增加到贫困生的录取中'.format(jobNum ,jobDict['feipin'])
                        print(msg)
                        break
        return enrolled,easy_list
        #返回的easylist是没有录取的非贫困生，

    else:
        print('没有非贫困生的名额')
        return None

def enroll_pinkun(jobInfo):
    print('------------------------------\n','接下来我们来进行贫困生的录取')
    easy_list = get_easyList(True)
    print('获得待录取资格的贫困生人数是 {0}'.format(len(easy_list)))
    enrolled = []
    count = 0
    for jobDict in jobInfo:
        n = jobDict.get('pinkun')
        if n != 0 and n != None:
            count = count + n
    if count:
        print('贫困生录取 名额 总数是 {0}'.format(count))
        for jobDict in jobInfo:  # 对招聘信息中招收贫困生的 岗位 进行遍历
            if jobDict.get('pinkun') != 0 and jobDict.get('pinkun') != None:
                n = jobDict['pinkun']
                jobNum = jobDict.get('num')
                for _ in range(n):  # 找几个人就循环多少次
                    bianhao = exeEnroll(jobNum, pin=True, easy_list=easy_list)
                    if bianhao != None:  # 录取了某个同学
                        print('  恭喜编号是{0}的同学(贫困生)被录取岗位{1}'.format(bianhao, jobNum))
                        # 操作录取的列表改变  改变jobinfo 增加录取的列表 减少带录取的easylist
                        enrollItem = {}
                        enrollItem[bianhao] = jobNum
                        enrolled.append(enrollItem)

                        jobDict['pinkun'] -= 1

                        i = 0
                        for appli in easy_list:
                            if appli[0] == bianhao:
                                del easy_list[i]
                            i += 1
                    else:
                        msg = '  对于岗位{0}，贫困生剩余{1}个名额'.format(jobNum, jobDict['pinkun'])
                        print(msg)
                        break
        return enrolled, easy_list
        # 返回的easylist是没有录取的非贫困生，

    else:
        print('没有非贫困生的名额')
        return None


def enrolled2full(enrolled_list):
    full_info = get_library(1)
    enrolled_full=[]
    for enroll_dict in enrolled_list:
        for bianhao in enroll_dict.keys(): #就一个，取出来编号
            for appli in full_info:
                if appli[0] == bianhao:
                    enroller = appli[:6]
                    enroller.append(enroll_dict[bianhao])
                    enrolled_full.append(enroller)
    return enrolled_full

def write_result(enrolled_feipin,enrolled_pinkun,job_Info,not_enrolled_feipin = None,not_enrolled_pinkun = None):
    result_book = Workbook()
    result_book.create_sheet('岗位剩余',0)
    result_book.create_sheet('贫困-录取的', 1)
    result_book.create_sheet('非贫困-录取的', 2)
    result_book.create_sheet('未录取的', 3)


    ws_jobInfo = result_book['岗位剩余']
    ws_jobInfo.cell(1,1).value = '岗位编号'
    ws_jobInfo.cell(1, 2).value = '贫困生剩余'
    ws_jobInfo.cell(1, 3).value = '非贫困生剩余'
    i = 2
    for job_dict in jobInfo:
        ws_jobInfo.cell(i,1).value = job_dict['num']
        ws_jobInfo.cell(i, 2).value = job_dict['pinkun']
        ws_jobInfo.cell(i, 3).value = job_dict['feipin']
        i+=1

    ws_pin_enrolled = result_book['贫困-录取的']
    ws_pin_enrolled.cell(1,1).value = '报名编号'
    ws_pin_enrolled.cell(1, 2).value = '姓名'
    ws_pin_enrolled.cell(1, 3).value = '性别'
    ws_pin_enrolled.cell(1, 4).value = '学号'
    ws_pin_enrolled.cell(1, 5).value = '学院'
    ws_pin_enrolled.cell(1, 6).value = '电话'
    ws_pin_enrolled.cell(1, 7).value = '录取岗位'
    pin_enrolled = enrolled2full(enrolled_pinkun)
    i = 2
    for enroller in pin_enrolled:
        for j in range(len(enroller)):
            ws_pin_enrolled.cell(i,j+1).value = enroller[j]
        i+=1

    ws_fpin_enrolled = result_book['非贫困-录取的']
    ws_fpin_enrolled.cell(1,1).value = '报名编号'
    ws_fpin_enrolled.cell(1, 2).value = '姓名'
    ws_fpin_enrolled.cell(1, 3).value = '性别'
    ws_fpin_enrolled.cell(1,4).value = '学号'
    ws_fpin_enrolled.cell(1, 5).value = '学院'
    ws_fpin_enrolled.cell(1, 6).value = '电话'
    ws_fpin_enrolled.cell(1, 7).value = '录取岗位'
    feipin_enrolled = enrolled2full(enrolled_feipin)
    i = 2
    for enroller in feipin_enrolled:
        for j in range(len(enroller)):
            ws_fpin_enrolled.cell(i,j+1).value = enroller[j]
        i+=1

    result_book.save(RESULT_PATH)


if __name__ == '__main__':
    print('Welcome to Tangzeling Working Automation Tool')
    print('现在的时间是： ',time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()),'\n','--------------------------------')

    jobInfo = get_library(2)
    enrolled_feipin,not_enrolled_feipin = enroll_feipin(jobInfo)
    enrolled_pinkun,not_enrolled_pinkun = enroll_pinkun(jobInfo)

    write_result(enrolled_feipin,enrolled_pinkun,jobInfo)
    print('录取结果在result.xlsx文件中 打开路径：',RESULT_PATH)
    input('这个不看了关了就行了')

