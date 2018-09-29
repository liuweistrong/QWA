from PATH import *
from openpyxl import Workbook
from openpyxl import load_workbook

#load library
def get_library(flag):
    appli_book = load_workbook(LIBRARY_PATH)
    library = []
    if flag == 1:  # 获取所有报名信息到列表
        ws = appli_book['library']
        max_row = ws.max_row
        #max col固定是14
        for row in range(2,max_row+1):
            appli_info = []
            for j in range(1,14+1):
                appli_info.append(ws.cell(row,j).value)
            library.append(appli_info)

            # [1, '姚佳莹', '女', 3017225028, '经管学院', 17302200916, '否', '否', '000001', None, None, '勤管办义务岗主管',None, None]
        return library

    if flag == 2:  #获取岗位需求信息到字典
        ws = appli_book['jobInfo']
        max_row = ws.max_row
        # max col固定是4
        for row in range(2, max_row + 1):
            job_info = {}
            job_info['num']=ws.cell(row, 1).value#编号
            job_info['pinkun']=ws.cell(row,3).value
            job_info['feipin']=ws.cell(row,4).value
            library.append(job_info)


            # class 'list'>: [{'num': '000001', 'pinkun': 3, 'feipin': 0}, {'num': '000002', 'pinkun': 2, 'feipin': 1},
            #                 {'num': '000004', 'pinkun': 4, 'feipin': 1}, {'num': '210101', 'pinkun': 1, 'feipin': None},
            #                 {'num': '280301', 'pinkun': 1, 'feipin': None},

        return library
