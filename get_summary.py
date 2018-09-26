import os
import time
import shutil
from openpyxl import Workbook
from openpyxl import load_workbook

#获取一个简历B48：G55区域的值，并形成一个列表
def get_appli_infolist(filename):
    appli_book = load_workbook(filename)
    appli_sheet = appli_book.active
    appli_info_list = []
    for i in range(48,56):
        for j in range(2,11):
            info_item = appli_sheet.cell(i,j)
            #if info_item.value != None:
            appli_info_list.append(info_item.value)
    print(appli_info_list)
    return appli_info_list
    # ['姓名', '张欢', None, None, '性别', '女', None, '照片               (仅纸质简历)', None, '学号', 3017213029, None, None, '学院',
    #  '药学院', None, None, None, '专业', '药学', None, None, '联系方式', 15032129628, None, None, None, '是否党员', '否', None, None,
    #  '是否贫困生', '是', None, None, None, '服从调剂', '是', None, None, '民族', '汉族', None, None, None, '应聘志愿', '志愿一', '岗位编号',
    #  '000001', '岗位名称', '勤管办 义务岗主管', None, None, None, None, '志愿二', '岗位编号', '000002', '岗位名称', '勤管办 兼职主管', None, None,
    #  None, None, '志愿三', '岗位编号', '230301', '岗位名称', '工会卫津路校区助理岗', None, None, None]


def is_valid(appli_info_list):
    if (appli_info_list[0] == '姓名' and
        appli_info_list[4] == '性别' and
        appli_info_list[9] == '学号' and
        appli_info_list[13] == '学院' and
        appli_info_list[22] == '联系方式' and
        appli_info_list[31] =='是否贫困生' and
        appli_info_list[36] == '服从调剂' and
        appli_info_list[46] == '志愿一' and
        appli_info_list[55] == '志愿二' and
        appli_info_list[64] == '志愿三'):
        return True
    else :
        return False

def applist2dict(appli_info_list):
    appli_dict = {}
    appli_dict[appli_info_list[0]] = appli_info_list[1]
    appli_dict[appli_info_list[4]] = appli_info_list[5]
    appli_dict[appli_info_list[9]] = appli_info_list[10]
    appli_dict[appli_info_list[13]] = appli_info_list[14]

    appli_dict[appli_info_list[22]] = appli_info_list[23]
    appli_dict[appli_info_list[31]] = appli_info_list[32]
    appli_dict[appli_info_list[36]] = appli_info_list[37]

    appli_dict['志愿一编号'] = appli_info_list[48]
    appli_dict['志愿一名称'] = appli_info_list[50]
    appli_dict['志愿二编号'] = appli_info_list[57]
    appli_dict['志愿二名称'] = appli_info_list[59]
    appli_dict['志愿三编号'] = appli_info_list[66]
    appli_dict['志愿三名称'] = appli_info_list[68]
    return appli_dict



def get_all_info():  #返回符合格式的信息，放到列表中【】，每个人一个字典。
    rootdir = r'D:\jianli\work'
    list = os.listdir(rootdir) #列出文件夹下所有的目录与文件
    print('检索到的所有文件',list)
    all_info_list=[]
    for i in range(0,len(list)):
        path = os.path.join(rootdir,list[i])
        if os.path.isfile(path):
               #你想对文件的操作
               appli_list = get_appli_infolist(path)
               if is_valid(appli_list):#如果和我们设置的检测关键字相同（简历没有修改位置），直接来处理
                    appi_dict = applist2dict(appli_list)
                    all_info_list.append(appi_dict)
               else:
                        #执行移动文件操作
                    if not os.path.exists('D:\jianli\work\FailToSum'):
                        os.makedirs('D:\jianli\work\FailToSum')
                    shutil.move(path,'D:\jianli\work\FailToSum')
    return all_info_list


def create_summary(all_info_list):
    #这里有个问题，保存不了save_path的内容
    #创建一个文件夹存放汇总的文件
    # rootdir = r'D:\jianli'
    operate_time = time.strftime("%Y-%m-%d+%H:%M", time.localtime())
    # dirname = rootdir + os.sep + operate_time + os.sep
    # if not os.path.exists('D:\jianli\汇总结果'):
    #     os.makedirs('D:\jianli\汇总结果')
    # os.makedirs(str(dirname))
    #operate_time_wb = time.strftime("%m-%d%H:%M", time.localtime())
    #wbname = operate_time_wb+'汇总.xlsx'
    #save_path = rootdir+ '\\'+ wbname
    wsname = '面试汇总'
    wb = Workbook()
    ws = wb.create_sheet(wsname, 0)
    i = 1
    for appli_dict in all_info_list:
        if i == 1:
            j = 1
            for val in appli_dict.keys():
                ws.cell(i, j).value = val
                j = j + 1
        else:
            j = 1
            for val in appli_dict.values():
                ws.cell(i,j).value = val
                j = j+1
        i = i + 1

    wb.save('D:\jianli\work\汇总简历.xlsx')
    if not os.path.exists('D:\jianli\work\SuccessToSum'):
        os.makedirs('D:\jianli\WORK\SuccessToSum')
    shutil.move('D:\jianli\work\汇总简历.xlsx', 'D:\jianli\WORK\SuccessToSum')

   # for appli_dict in all_info_list:
    #    for value in appli_dict.values():



if __name__ == '__main__':
    list = get_all_info()
    print(list)
    create_summary(list)
