from PATH import *
from openpyxl import Workbook
from transform import enrolled2full
from load import get_library
from openpyxl import load_workbook
def write_result(enrolled_feipin,enrolled_pinkun,job_Info,not_enrolled_feipin = None,not_enrolled_pinkun = None):
    result_book = Workbook()
    result_book.create_sheet('岗位剩余',0)
    result_book.create_sheet('贫困-录取的', 1)
    result_book.create_sheet('非贫困-录取的', 2)
    result_book.create_sheet('未录取的', 3)


    ws_jobInfo = result_book['岗位剩余']
    ws_jobInfo.cell(1,1).value = '岗位编号'
    ws_jobInfo.cell(1,2).value = '岗位名称'
    ws_jobInfo.cell(1, 3).value = '贫困生剩余'
    ws_jobInfo.cell(1, 4).value = '非贫困生剩余'
    i = 2
    for job_dict in job_Info:
        ws_jobInfo.cell(i,1).value = job_dict['num']
        ws_jobInfo.cell(i, 2).value = job_dict['name']
        ws_jobInfo.cell(i, 3).value = job_dict['pinkun']
        ws_jobInfo.cell(i, 4).value = 0
        i+=1

    ws_pin_enrolled = result_book['贫困-录取的']
    ws_pin_enrolled.cell(1,1).value = '报名编号'
    ws_pin_enrolled.cell(1, 2).value = '姓名'
    ws_pin_enrolled.cell(1, 3).value = '性别'
    ws_pin_enrolled.cell(1, 4).value = '学号'
    ws_pin_enrolled.cell(1, 5).value = '学院'
    ws_pin_enrolled.cell(1, 6).value = '电话'
    ws_pin_enrolled.cell(1, 7).value = '录取岗位'
    pin_enrolled = enrolled2full(enrolled_pinkun)
    i = 2
    for enroller in pin_enrolled:
        for j in range(len(enroller)):
            ws_pin_enrolled.cell(i,j+1).value = enroller[j]
        ws_pin_enrolled.cell(i, 8).value = get_job_name(enroller[6])
        i+=1

    ws_fpin_enrolled = result_book['非贫困-录取的']
    ws_fpin_enrolled.cell(1,1).value = '报名编号'
    ws_fpin_enrolled.cell(1, 2).value = '姓名'
    ws_fpin_enrolled.cell(1, 3).value = '性别'
    ws_fpin_enrolled.cell(1,4).value = '学号'
    ws_fpin_enrolled.cell(1, 5).value = '学院'
    ws_fpin_enrolled.cell(1, 6).value = '电话'
    ws_fpin_enrolled.cell(1, 7).value = '录取岗位'
    feipin_enrolled = enrolled2full(enrolled_feipin)
    i = 2
    for enroller in feipin_enrolled:
        for j in range(len(enroller)):
            ws_fpin_enrolled.cell(i,j+1).value = enroller[j]
        ws_fpin_enrolled.cell(i, 8).value = get_job_name(enroller[6])
        i+=1

    result_book.save(RESULT_PATH)


def get_job_name(num):
    jobInfo_book = load_workbook(LIBRARY_PATH)
    job_Info = jobInfo_book['jobInfo']
    max_row = job_Info.max_row
    for raw in range(2,max_row+1):
        if(num == job_Info.cell(raw,1).value):
            return job_Info.cell(raw,2).value

